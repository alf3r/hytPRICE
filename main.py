import logging
import tkinter as tk
from datetime import datetime
from time import gmtime, strftime
from tkinter import filedialog, messagebox

import openpyxl.utils
from spire.doc.common import *

import hytALGORITHMS

logger = logging.getLogger(__name__)


########################################################################################################################
# Основное тело программы
class HytesPrices(tk.Tk):
    def __del__(self):
        logger.info('выключение программы')
    def __init__(self):
        super().__init__()
        self.Btn12 = None
        self.Btn5 = None
        self.Btn14 = None
        self.Btn13 = None
        self.myTimestamp = None
        self.Btn11 = None
        self.Btn4 = None
        self.Btn3 = None
        self.Btn2 = None
        self.Btn1 = None
        self.Btn8 = None
        self.Btn7 = None
        self.Btn9 = None
        self.Btn10 = None
        self.Btn6 = None
        self.myEntry3 = None
        self.myEntry2 = None
        self.myEntry1 = None
        self.myEntryCurrency = None
        self.myLabel6 = None
        self.myLabel5 = None
        self.myLabel1 = None
        self.myLabelCurrency = None
        self.LabelFrameTest = None
        self.LabelFrameRUS = None
        self.LabelFrameCPQ = None
        self.myButtons = None
        self.LabelFrameBasic = None
        self.title("Hytrus v1.5")
        # self.geometry("450x400")

        self.dirCommon = os.getcwd() + '/workDIR/'
        self.dirLog = os.getcwd() + '/workDIR/99_logs/'
        if not os.path.exists(self.dirCommon):
            os.makedirs(self.dirCommon)
        if not os.path.exists(self.dirLog):
            os.makedirs(self.dirLog)
        self.strComplete = 'Успех! - '

        # define variables
        self.myPrefix = tk.StringVar(value='')
        self.price_date = tk.StringVar(value='2026.Q3')
        self.price_version = tk.StringVar(value='V1.0')
        self.currentCurrency = tk.StringVar(value='7.16')
        self.isAntenna = tk.IntVar(value=0)
        self.isOnlyAntenna = tk.IntVar(value=0)
        #self.isAntenna.trace_add("write", self.checkbutton_changed)

        now = datetime.now()
        logFilename = self.dirLog + 'Hytrus_' + now.strftime("%Y-%m-%d") + '.log'
        logging.basicConfig(level=logging.INFO, filename=logFilename,
                            format="%(asctime)s %(levelname)s %(message)s")
        self.create_widgets()

    def create_widgets(self):
        # кнопки
        self.LabelFrameBasic = tk.LabelFrame(self, text="Исходные данные", relief='groove', font='Helvetica 12 bold')
        self.LabelFrameBasic.grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky='WEW')
        self.LabelFrameCPQ = tk.LabelFrame(self, text="Обработка прайсов из CPQ", relief='groove', font='Helvetica 12 bold')
        self.LabelFrameCPQ.grid(row=1, column=0, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.LabelFrameRUS = tk.LabelFrame(self, text="Обработка RUS прайсов", relief='groove', font='Helvetica 12 bold')
        self.LabelFrameRUS.grid(row=1, column=1, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.LabelFrameTest = tk.LabelFrame(self, text="тестовые", relief='groove', font='Helvetica 10 bold')
        self.LabelFrameTest.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky='WEW')
        # надписи
        self.myLabel1 = tk.Label(self.LabelFrameBasic, text='Введите название прайса')
        self.myLabel5 = tk.Label(self.LabelFrameBasic, text='Введите дату прайса')
        self.myLabel6 = tk.Label(self.LabelFrameBasic, text='Введите версию прайса')
        self.myLabelCurrency = tk.Label(self.LabelFrameBasic, text='Введите курс USD-CNY')
        # self.myLabel2 = tk.Label(self, text='тестовые', font='Helvetica 10')
        # self.myLabel3 = tk.Label(self, text='Обработка прайсов из CPQ', font='Helvetica 10 bold')
        # self.myLabel4 = tk.Label(self, text='Обработка RUS прайсов', font='Helvetica 10 bold')
        # поле ввода префикса
        self.myEntry1 = tk.Entry(self.LabelFrameBasic, textvariable=self.myPrefix)
        self.myEntry2 = tk.Entry(self.LabelFrameBasic, textvariable=self.price_date)
        self.myEntry3 = tk.Entry(self.LabelFrameBasic, textvariable=self.price_version)
        self.myEntryCurrency = tk.Entry(self.LabelFrameBasic, textvariable=self.currentCurrency)
        # чек бокс Антенны
        #self.myCheckButton1 = tk.Checkbutton(text="Антенны", variable=self.isAntenna)
        #self.myCheckButton1 = tk.Checkbutton(text="Включить антенны в прайс?", variable=self.isOnlyAntenna)
        # кнопки
        self.Btn6 = tk.Button(self.LabelFrameCPQ,text="Make RUS Price from Folder", command=self.clickButton_rusPricesFromFolder)
        self.Btn10 = tk.Button(self.LabelFrameCPQ,text="Make RUS Antenna price from Folder", command=self.clickButton_rusAntennaFromFolder)
        self.Btn9 = tk.Button(self.LabelFrameCPQ,text="Convert SYSTEM_CPQ prices to terminals type", command=self.clickButton_prepareSystem)
        self.Btn13 = tk.Button(self.LabelFrameCPQ, text="Convert SYSTEM_CPQ prices to SYSTEM_RUS type", command=self.clickButton_convertSystemCNY)

        self.Btn7 = tk.Button(self.LabelFrameRUS, text="Make DOCX from Excel", command=self.clickButton_makeDOCx)
        self.Btn8 = tk.Button(self.LabelFrameRUS, text="compare 2 Excel RUS", command=self.clickButton_compareEXCELrus)
        self.Btn14 = tk.Button(self.LabelFrameRUS, text="Combine SYSTEM prices from folder", command=self.clickButton_cpqSystemCombine)

        self.Btn1 = tk.Button(self.LabelFrameTest,text="Combine TERMINAL prices from folder", command=self.clickButton_cpqPricesCombine)
        self.Btn2 = tk.Button(self.LabelFrameTest,text="Combine sheets on only one sheet", command=self.clickButton_onePagePriceUnique)
        self.Btn3 = tk.Button(self.LabelFrameTest,text="Make Rus Price from File", command=self.clickButton_rusPricesFromFile)
        self.Btn4 = tk.Button(self.LabelFrameTest,text="Divide RUS One-page-price by Sheets", command=self.clickButton_dividePriceBySheets)
        self.Btn11 = tk.Button(self.LabelFrameTest,text="Find new BOMs", command=self.clickButton_findNewBOM)
        self.Btn5 = tk.Button(self.LabelFrameTest,text="Get Images From Hytes price", command=self.clickButton_getImagesFromExcel)
        self.Btn12 = tk.Button(self.LabelFrameTest,text="Make РРЦ template", command=self.clickButton_convertGPL2RRC)

        #отрисовка
        self.myLabel1.grid(row=0, column=0, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.myEntry1.grid(row=1, column=0, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.myLabel5.grid(row=0, column=1, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.myEntry2.grid(row=1, column=1, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.myLabel6.grid(row=0, column=2, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.myEntry3.grid(row=1, column=2, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.myLabelCurrency.grid(row=0, column=3, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.myEntryCurrency.grid(row=1, column=3, columnspan=1, padx=5, pady=5, sticky='WEW')

        #Обработка прайсов из CPQ
        self.Btn6.grid(row=0, column=0, columnspan=1, padx=5, pady=5, sticky='WEW', )
        self.Btn10.grid(row=1, column=0, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.Btn9.grid(row=2, column=0, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.Btn13.grid(row=3, column=0, columnspan=1, padx=5, pady=5, sticky='WEW')

        self.Btn7.grid(row=0, column=0, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.Btn8.grid(row=1, column=0, columnspan=1, padx=5, pady=5, sticky='WEW')
        self.Btn14.grid(row=2, column=0, columnspan=1, padx=5, pady=5, sticky='WEW')

        #Тестовые
        self.Btn12.grid(row=0, column=1, columnspan=1, padx=1, pady=1, sticky='WEW')
        self.Btn5.grid(row=1, column=1, columnspan=1, padx=1, pady=1, sticky='WEW')
        self.Btn1.grid(row=2, column=1, columnspan=1, padx=1, pady=1, sticky='WEW')
        self.Btn2.grid(row=0, column=2, columnspan=1, padx=1, pady=1, sticky='WEW')
        self.Btn3.grid(row=1, column=2, columnspan=1, padx=1, pady=1, sticky='WEW')
        self.Btn4.grid(row=2, column=2, columnspan=1, padx=1, pady=1, sticky='WEW')
        self.Btn11.grid(row=0, column=3, columnspan=1, padx=1, pady=1, sticky='WEW')

        logger.info('----------------')
        logger.info('ЗАПУСК ПРОГРАММЫ')


    #6 КЛИК - Сделать прайс русский из папки
    def clickButton_rusPricesFromFolder(self):
        logger.info('-----Начат "Сделать прайс русский из папки "')
        self.isOnlyAntenna = tk.IntVar(value=0)
        dirIn = filedialog.askdirectory(initialdir=self.dirCommon,
                                        title='Выберите папку с терминальными прайсами из CPQ')
        HYTES_PriceList_Filename = filedialog.askopenfilename(initialdir=self.dirCommon,
                                                              initialfile="Прайс Хайтес.xlsx",
                                                              defaultextension=".xlsx",
                                                              title='Выберите "Прайс Хайтес"')
        kwargs1 = {
            "dirIn": dirIn
        }
        cpqPricesCombine_Filename = self.clickButton_cpqPricesCombine(**kwargs1)
        kwargs2 = {
            "cpqPricesCombine_Filename": cpqPricesCombine_Filename,
            "HYTES_PriceList_Filename": HYTES_PriceList_Filename,
        }
        self.clickButton_rusPricesFromFile(**kwargs2)
        numNewBOM = hytALGORITHMS.findNewBOM(cpqPricesCombine_Filename)
        print("Новых BOM-кодов: " + str(numNewBOM))
        try:
            os.remove(cpqPricesCombine_Filename)
        except OSError:
            pass
        print(self.strComplete + self.clickButton_rusPricesFromFolder.__name__)
        logger.info('Закончен "Сделать прайс русский из папки ". Новых BOM-кодов: ' + str(numNewBOM))
        messagebox.showinfo(title='Успех!', message='Сделан русский прайс EXCEL для терминалов из папки!')


    #10 КЛИК - Сделать прайс Антенны русский из папки
    def clickButton_rusAntennaFromFolder(self):
        logger.info('-----Начат "Сделать прайс Антенны русский из папки"')
        self.isOnlyAntenna = tk.IntVar(value=1)
        dirIn = filedialog.askdirectory(initialdir=self.dirCommon,
                                        title='Выберите папку с терминальными прайсами из CPQ')
        HYTES_PriceList_Filename = filedialog.askopenfilename(initialdir=self.dirCommon,
                                                              initialfile="Прайс Хайтес.xlsx",
                                                              defaultextension=".xlsx",
                                                              title='Выберите "Прайс Хайтес"')
        kwargs1 = {
            "dirIn": dirIn
        }
        cpqPricesCombine_Filename = self.clickButton_cpqPricesCombine(**kwargs1)

        kwargs2 = {
            "cpqPricesCombine_Filename": cpqPricesCombine_Filename,
            "HYTES_PriceList_Filename": HYTES_PriceList_Filename,
        }
        rusPricesFromFile_Filename = self.clickButton_rusPricesFromFile(**kwargs2)

        kwargs3 = {
            "rusPricesFromFile_Filename": rusPricesFromFile_Filename
        }
        onePagePriceUnique_filename = self.clickButton_onePagePriceUnique(**kwargs3)

        kwargs4 = {
            "onePagePriceUnique_filename": onePagePriceUnique_filename
        }
        self.clickButton_dividePriceBySheets(**kwargs4)
        print(self.strComplete + self.clickButton_rusAntennaFromFolder.__name__)
        logger.info('Закончен "Сделать прайс Антенны русский из папки"')

        numNewBOM = hytALGORITHMS.findNewBOM(cpqPricesCombine_Filename)
        print("Новых BOM-кодов: " + str(numNewBOM))
        try:
            # print(cpqPricesCombine_Filename)
            # print(rusPricesFromFile_Filename)
            # print(onePagePriceUnique_filename)
            os.remove(cpqPricesCombine_Filename)
            os.remove(rusPricesFromFile_Filename)
            os.remove(onePagePriceUnique_filename)
        except OSError:
            pass
        messagebox.showinfo(title='Успех!', message='Сделан русский прайс EXCEL для Антенн из папки!')

    #3 КЛИК - Подставить цены из прайса Хайтес
    def clickButton_rusPricesFromFile(self, **kwargs):
        logger.info('-----Начат "Подставить цены из прайса Хайтес"')
        self.myTimestamp = strftime("%Y%m%d%H%M%S", gmtime())
        HYTES_PriceList_Filename = kwargs.get("HYTES_PriceList_Filename")
        if HYTES_PriceList_Filename:
            pass
            # print("clickButton_rusPricesFromFile HYTES_PriceList_Filename ---->>> " + HYTES_PriceList_Filename)
        else:
            HYTES_PriceList_Filename = filedialog.askopenfilename(initialdir=self.dirCommon, title='Выберите "Прайс Хайтес"')
        cpqPricesCombine_Filename = kwargs.get("cpqPricesCombine_Filename")
        if cpqPricesCombine_Filename:
            pass
            # print("clickButton_rusPricesFromFile cpqPricesCombine_Filename ---->>> " + cpqPricesCombine_Filename)
        else:
            cpqPricesCombine_Filename = filedialog.askopenfilename(title='Выберите файл "Combine"')
        rusPrice_Filename = hytALGORITHMS.setRusPrices(HYTES_PriceList_Filename, cpqPricesCombine_Filename)
        print(self.strComplete + self.clickButton_rusPricesFromFile.__name__)
        logger.info('Закончен "Подставить цены из прайса Хайтес". Создан файл: ' + str(rusPrice_Filename))
        return rusPrice_Filename


    #1 КЛИК - Комбинировать ТЕРМИНАЛЬНЫЕ прайсы в одну книгу в виде закладок
    def clickButton_cpqPricesCombine(self, **kwargs):
        logger.info(
            '-----Начат "Комбинировать ТЕРМИНАЛЬНЫЕ прайсы в одну книгу в виде закладок"')
        myTimestamp = strftime("%Y%m%d%H%M%S", gmtime())
        dirIn = kwargs.get("dirIn")
        if dirIn:
            pass
            # print("clickButton_cpqPricesCombine dirIn ---->>> " + dirIn)
        else:
            dirIn = filedialog.askdirectory(initialdir=self.dirCommon, title='Выберите папку с терминальными прайсами из CPQ')
        os.chdir(dirIn)
        dirName = os.path.basename(dirIn)
        self.myPrefix = dirName
        self.myEntry1.delete(0, 'end')
        self.myEntry1.insert(0, dirName)
        logger.info('Название папки: "' + dirName + '"')
        wbOut = openpyxl.Workbook()
        # берем перечень файлов с прайсами
        filesPrices = sorted(os.listdir(dirIn))
        # собираем прайсы в одну книгу #####################################################################################
        errors = hytALGORITHMS.cpqPricesCombine(filesPrices, wbOut)
        ####################################################################################################################
        # сохраняем новую комбинированную книгу
        #dirOut = filedialog.askdirectory(title='Выберите папку, в которую сохранить общий прайс')
        dirOut = self.dirCommon
        os.chdir(dirOut)
        outputFileName: str = dirOut + str(self.myPrefix) + '-' + myTimestamp + '.xlsx'
        print(outputFileName)
        del wbOut['Sheet']
        wbOut.save(outputFileName)
        wbOut.close()
        print(self.strComplete +  self.clickButton_cpqPricesCombine.__name__)
        for err in errors:
            if err == '[]' or err == '' or err == ' ':
                continue
            else:
                logger.info(err)
        logger.info('Закончен "Комбинировать ТЕРМИНАЛЬНЫЕ прайсы в одну книгу в виде закладок". Создан файл: ' + outputFileName)
        return outputFileName


    #2 КЛИК - Сделать общий прайс на одной вкладке с уникальными позициями
    def clickButton_onePagePriceUnique(self, **kwargs):
        logger.info(
            '-----Начат "Сделать общий прайс на одной вкладке с уникальными позициями"')
        cpqPricesCombine_Filename = kwargs.get("rusPricesFromFile_Filename")
        if cpqPricesCombine_Filename:
            pass
            # print("clickButton_cpqPricesCombine cpqPricesCombine_Filename ---->>> " + cpqPricesCombine_Filename)
        else:
            cpqPricesCombine_Filename = filedialog.askopenfilename(initialdir=self.dirCommon, title='Выберите файл "RUS"')
        idxXLSX = cpqPricesCombine_Filename.find('.xlsx')
        priceOnepage_filename = cpqPricesCombine_Filename[:idxXLSX] + '-OnePageUnique' + cpqPricesCombine_Filename[idxXLSX:]
        print(priceOnepage_filename)
        #########################################################################################
        hytALGORITHMS.onePagePrice(cpqPricesCombine_Filename, priceOnepage_filename, self.isOnlyAntenna.get())
        ################################################################################################################
        print(self.strComplete +  self.clickButton_onePagePriceUnique.__name__)
        logger.info('Закончен "Сделать общий прайс на одной вкладке с уникальными позициями". Имя файла: ' + priceOnepage_filename)
        return priceOnepage_filename

    #5 КЛИК - Сохранить изображения из прайса Хайтес
    def clickButton_getImagesFromExcel(self):
        logger.info('-----Начат "Сохранить изображения из прайса Хайтес"')
        self.myTimestamp = strftime("%Y%m%d%H%M%S", gmtime())
        HYTES_PriceList_Filename = filedialog.askopenfilename(initialdir=self.dirCommon, title='Выберите "Прайс Хайтес"')
        print('Извлекаю изображения из прайса...')
        hytALGORITHMS.saveImagesExcel2PNG(HYTES_PriceList_Filename)
        logger.info('Закончен "Сохранить изображения из прайса Хайтес"')
        print(self.strComplete + self.clickButton_getImagesFromExcel.__name__)


    #6 КЛИК - Разделить один лист на много листов
    def clickButton_dividePriceBySheets(self, **kwargs):
        logger.info('-----Начат "Разделить один лист на много листов"')
        self.myTimestamp = strftime("%Y%m%d%H%M%S", gmtime())
        onePagePriceUnique_filename = kwargs.get("onePagePriceUnique_filename")
        if onePagePriceUnique_filename:
            pass
            # print("clickButton_dividePriceBySheets Combined_Price_Filename ---->>> " + Combined_Price_Filename)
        else:
            onePagePriceUnique_filename = filedialog.askopenfilename(initialdir=self.dirCommon, title='Выберите "OnePageUnique" файл')
        ################################################################################################################
        dividePriceBySheets_filename = hytALGORITHMS.dividePriceBySheets(onePagePriceUnique_filename)
        ################################################################################################################
        print(self.strComplete + self.clickButton_getImagesFromExcel.__name__)
        logger.info('Закончен "Разделить один лист на много листов". Имя файла: ' + dividePriceBySheets_filename)
        return dividePriceBySheets_filename

    #7 КЛИК - Найти новые BOM-коды
    def clickButton_findNewBOM(self):
        logger.info('-----Начат "Найти новые BOM-коды"')
        self.myTimestamp = strftime("%Y%m%d%H%M%S", gmtime())
        priceCombine_filename = filedialog.askopenfilename(initialdir=self.dirCommon, title='Выберите файл "Combine"')
        hytALGORITHMS.findNewBOM(priceCombine_filename)
        print(self.strComplete + self.clickButton_findNewBOM.__name__)
        logger.info('Закончен "Найти новые BOM-коды"')

    #6 КЛИК - Преобразовать Excel в Word
    def clickButton_makeDOCx(self):
        logger.info('-----Начат "Преобразовать Excel в Word"')
        priceDate = self.price_date.get()
        priceVersion = self.price_version.get()
        excel_File_Rus = filedialog.askopenfilename(initialdir=self.dirCommon, title='Выберите файл .XLSX с прайсом "RUS"')
        word_template = filedialog.askopenfilename(initialdir=self.dirCommon + '01_templates',
                                                   title='Выберите файл с шаблоном прайса DOCX')
        word_files = hytALGORITHMS.convertGPL2RRC(word_template, priceVersion, priceDate)
        # templateRRC = filedialog.askopenfilename(title='Выберите файл с шаблоном DOCX РРЦ')
        dirImage = filedialog.askdirectory(initialdir=self.dirCommon + '00_Images',
                                           title='Выберите папку с изображениями')
        errors = hytALGORITHMS.excel2word(excel_File_Rus, word_files[0], word_files[1], dirImage, priceDate, priceVersion)
        for err in errors:
            if err == [] or err == '' or err == ' ':
                continue
            else:
                logger.info(err)
        try:
            os.remove(word_files[0])
            os.remove(word_files[1])
        except OSError:
            pass
        logger.info('Закончен "Преобразовать Excel в Word"')
        print(self.strComplete + self.clickButton_makeDOCx.__name__)
        messagebox.showinfo(title='Успех!', message='Преобразован прайс Excel в Word (GPL\РРЦ)!')

    #7 КЛИК - сравнение двух файлов EXCEL RUS
    def clickButton_compareEXCELrus(self):
        logger.info('-----Начат "Сравнение двух файлов EXCEL RUS"')
        OldFileName = filedialog.askopenfilename(title='Выберите СТАРЫЙ прайс "RUS" для сравнения')
        NewFileName = filedialog.askopenfilename(initialdir=self.dirCommon, title='Выберите НОВЫЙ прайс "RUS" для сравнения')
        print('Сравниваю...')
        errors = hytALGORITHMS.compareEXLSrus(NewFileName, OldFileName)
        print(self.strComplete + self.clickButton_compareEXCELrus.__name__)
        for err in errors:
            if err == '[]' or err == '' or err == ' ':
                continue
            else:
                logger.info(err)
        logger.info('Закончен "Сравнение двух файлов EXCEL RUS"')
        messagebox.showinfo(title='Успех!', message='Завершено сравнение двух прайсов EXCEL!')

   #9 КЛИК - Конвертировать системный прайс в терминальный
    def clickButton_prepareSystem(self):
        logger.info('-----Начат "Конвертировать системный прайс в терминальный"')
        dirIn = filedialog.askdirectory(initialdir=self.dirCommon, title='Выберите папку с СИСТЕМНЫМИ прайсами из CPQ')
        os.chdir(dirIn)
        # берем перечень файлов с прайсами
        filesPrices = sorted(os.listdir(dirIn))
        for n, item in enumerate(filesPrices):
            os.rename(item, 'sys_' + item)
        filesPrices = sorted(os.listdir(dirIn))
        for n, item in enumerate(filesPrices):
            wbIn = openpyxl.load_workbook(item)
            try:
                sheetInName = os.path.basename(item)
                sheetInName = sheetInName.replace(" ", "")
                sheetInName = sheetInName.replace("\xa0", "")
                sheetInName = sheetInName.replace("-", "")
                sheetInName = sheetInName.replace("sys_", "")
                sheetInName = sheetInName.replace(".xlsx", "")
                sheetIn = wbIn['Sheet0']
                sheetOut = wbIn.create_sheet(sheetInName)
                #############################################################################конвертируем лист системный
                hytALGORITHMS.prepare_SYSTEMsheet(sheetIn, sheetOut)
                ########################################################################################################
                # сохраняем новые конвертированные книги
                outputFileName: str = dirIn + '/' + sheetInName + '.xlsx'
                del wbIn['Sheet0']
                wbIn.save(outputFileName)
                print(outputFileName)
            finally:
                wbIn.close()
        print(self.strComplete + self.clickButton_prepareSystem.__name__)
        logger.info('Закончен "Конвертировать прайс CPQ в терминальный"')
        messagebox.showinfo(title='Успех!', message='Завершено конвертирование прайс CPQ в терминальный!')

    #10 КЛИК - Конвертировать GPL шаблон в РРЦ
    def clickButton_convertGPL2RRC(self):
        priceDate = self.price_date.get()
        priceVersion = self.price_version.get()
        logger.info('-----Начат "Конвертировать GPL шаблон в РРЦ"')
        word_template = filedialog.askopenfilename(initialdir=self.dirCommon, title='Выберите файл с шаблоном прайса DOCX')
        hytALGORITHMS.convertGPL2RRC(word_template, priceVersion, priceDate)
        print(self.strComplete + self.clickButton_convertGPL2RRC.__name__)
        logger.info('Закончен "Конвертировать GPL шаблон в РРЦ"')

    # 11 КЛИК - Конвертировать системный прайс в российский CNY
    def clickButton_convertSystemCNY(self):
        logger.info('-----Начат "Конвертировать системный прайс в российский CNY"')
        myTimestamp = strftime("%Y%m%d%H%M%S", gmtime())
        myCurrency = self.currentCurrency.get()
        dirIn = filedialog.askdirectory(initialdir=self.dirCommon, title='Выберите папку с СИСТЕМНЫМИ прайсами из CPQ')
        os.chdir(dirIn)
        # берем перечень файлов с прайсами
        filesPrices = sorted(os.listdir(dirIn))
        for n, item in enumerate(filesPrices):
            wbIn = openpyxl.load_workbook(item)
            sheetIn = wbIn['Sheet0']
            try:
                sheetInName = os.path.basename(item)
                sheetInName = sheetInName.replace(" ", "")
                sheetInName = sheetInName.replace("\xa0", "")
                sheetInName = sheetInName.replace("-", "")
                sheetInName = sheetInName.replace("sys_", "")
                sheetInName = sheetInName.replace(".xlsx", "")
                sheetOut = wbIn.create_sheet(sheetInName)
                #############################################################################конвертируем лист системный
                hytALGORITHMS.sheetCPQ2RUS(sheetIn, sheetOut, myCurrency)
                ########################################################################################################
                # сохраняем новые конвертированные книги
                outputFileName: str = dirIn + '/' + sheetInName + '-' + myTimestamp + '.xlsx'
                del wbIn['Sheet0']
                wbIn.save(outputFileName)
                print(outputFileName)
            finally:
                wbIn.close()
        print(self.strComplete + self.clickButton_convertSystemCNY.__name__)
        logger.info('Закончен "Конвертировать прайс CPQ в системный российский CNY"')
        messagebox.showinfo(title='Успех!', message='Завершено конвертирование прайса CPQ в системный российский CNY!')


    #12 КЛИК - Комбинировать СИСТЕМНЫЕ прайсы в одну книгу в виде закладок
    def clickButton_cpqSystemCombine(self, **kwargs):
        logger.info(
            '-----Начат "Комбинировать СИСТЕМНЫЕ прайсы в одну книгу в виде закладок"')
        myTimestamp = strftime("%Y%m%d%H%M%S", gmtime())
        dirIn = kwargs.get("dirIn")
        if dirIn:
            pass
            # print("clickButton_cpqPricesCombine dirIn ---->>> " + dirIn)
        else:
            dirIn = filedialog.askdirectory(initialdir=self.dirCommon, title='Выберите папку с терминальными прайсами из CPQ')
        os.chdir(dirIn)
        dirName = os.path.basename(dirIn)
        self.myPrefix = dirName
        self.myEntry1.insert(0, dirName)
        logger.info('Название папки: "' + dirName + '"')
        wbOut = openpyxl.Workbook()
        # берем перечень файлов с прайсами
        filesPrices = sorted(os.listdir(dirIn))
        # собираем прайсы в одну книгу #####################################################################################
        hytALGORITHMS.cpqSystemCombine(filesPrices, wbOut)
        ####################################################################################################################
        # сохраняем новую комбинированную книгу
        #dirOut = filedialog.askdirectory(title='Выберите папку, в которую сохранить общий прайс')
        dirOut = self.dirCommon
        os.chdir(dirOut)
        outputFileName: str = dirOut + str(self.myPrefix) + '-' + myTimestamp + '.xlsx'
        print(outputFileName)
        wbOut.save(outputFileName)
        wbOut.close()
        print(self.strComplete +  self.clickButton_cpqPricesCombine.__name__)
        logger.info('Закончен "Комбинировать СИСТЕМНЫЕ прайсы в одну книгу в виде закладок". Создан файл: ' + outputFileName)
        return outputFileName

########################################################################################################################
########################################################################################################################
if __name__ == "__main__":
  app = HytesPrices()
  app.mainloop()